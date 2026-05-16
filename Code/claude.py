from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── style helpers ─────────────────────────────────────────────────────────────
BLUE_INPUT  = "0000FF"   # hardcoded user inputs
BLACK_CALC  = "000000"   # formulas
WHITE       = "FFFFFF"

def fill(hex_color): return PatternFill("solid", start_color=hex_color)
def thin_border(top=False, bot=False, left=False, right=False):
    s = Side(style="thin")
    return Border(
        top=s if top else None, bottom=s if bot else None,
        left=s if left else None, right=s if right else None)

def all_border():
    s = Side(style="thin")
    return Border(top=s, bottom=s, left=s, right=s)

def thick_border():
    s = Side(style="medium")
    return Border(top=s, bottom=s, left=s, right=s)

def style(cell, value=None, bold=False, color="000000", bg=None,
          align="center", sz=10, italic=False, fmt=None, border=None):
    if value is not None:
        cell.value = value
    cell.font = Font(name="Arial", bold=bold, color=color, size=sz, italic=italic)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bg: cell.fill = fill(bg)
    if fmt: cell.number_format = fmt
    if border: cell.border = border
    return cell

def header(ws, row, col, text, bg="1F4E79", fg="FFFFFF", colspan=1, sz=10):
    c = ws.cell(row=row, column=col)
    style(c, text, bold=True, color=fg, bg=bg, sz=sz)
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+colspan-1)
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 20)
    return c

def section_label(ws, row, col, text, bg="2E75B6", colspan=1):
    c = ws.cell(row=row, column=col)
    style(c, text, bold=True, color=WHITE, bg=bg, sz=10)
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+colspan-1)
    ws.row_dimensions[row].height = 22
    return c

def input_cell(ws, row, col, value, fmt="0.00"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", color=BLUE_INPUT, bold=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = fill("EBF3FB")
    if fmt: c.number_format = fmt
    c.border = all_border()
    return c

def formula_cell(ws, row, col, formula, fmt="0.00000E+00"):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = Font(name="Arial", color=BLACK_CALC, size=10)
    c.alignment = Alignment(horizontal="right", vertical="center")
    if fmt: c.number_format = fmt
    return c

def label_cell(ws, row, col, text, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Arial", bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    return c

def row_bg(ws, row, col_start, col_end, hex_color):
    for c in range(col_start, col_end+1):
        ws.cell(row, c).fill = fill(hex_color)

def make_chart(ws, title, x_col, y_cols, series_titles,
               top_left, data_start, n_rows, colors=None, w=22, h=14):
    if colors is None:
        colors = ["FF0000","0070C0","00B050","FFC000","7030A0","FF6600"]
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "gz (mGal)"
    chart.x_axis.title = "Posisi x (m)"
    chart.y_axis.numFmt = "0.00E+00"
    chart.width = w; chart.height = h
    last = data_start + n_rows - 1
    cats = Reference(ws, min_col=x_col, min_row=data_start, max_row=last)
    for i, (col, stitle) in enumerate(zip(y_cols, series_titles)):
        data = Reference(ws, min_col=col, min_row=data_start, max_row=last)
        chart.add_data(data)
        s = chart.series[-1]
        s.title = SeriesLabel(v=stitle)
        s.graphicalProperties.line.solidFill = colors[i % len(colors)]
        s.graphicalProperties.line.width = 20000
    chart.set_categories(cats)
    ws.add_chart(chart, top_left)

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 2  —  MODEL BOLA
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.active
ws2.title = "Sheet2_ModelBola"
ws2.sheet_view.showGridLines = True
ws2.freeze_panes = "B27"

# column widths
col_w2 = {"A":6,"B":12,"C":12,"D":12,"E":12,"F":12,"G":12,
           "H":4,"I":12,"J":12,"K":12,"L":12,"M":12,
           "N":4,"O":12,"P":12,"Q":12,"R":12,"S":12}
for c,w in col_w2.items(): ws2.column_dimensions[c].width = w

# ── ROW 1: main title ─────────────────────────────────────────────────────────
ws2.merge_cells("A1:S1")
style(ws2["A1"], "MODEL BOLA — Pemodelan Gravitasi 2D",
      bold=True, color=WHITE, bg="1F4E79", sz=14)
ws2.row_dimensions[1].height = 30

# ── ROW 2-3: constant parameters ─────────────────────────────────────────────
ws2.merge_cells("A2:A3"); style(ws2["A2"], "Parameter\nTetap", bold=True, bg="D6E4F0", sz=9)
label_cell(ws2, 2, 2, "G (mGal konst.)", bold=True, align="right")
c = ws2.cell(2, 3, 6.674e-6); c.font=Font(name="Arial",color=BLUE_INPUT,bold=True,size=10)
c.number_format="0.000E+00"; c.alignment=Alignment(horizontal="center",vertical="center")
c.fill=fill("EBF3FB"); c.border=all_border()
G_CELL = "$C$2"

label_cell(ws2, 2, 4, "x-start (m)", bold=True, align="right")
input_cell(ws2, 2, 5, 0, "0")
label_cell(ws2, 2, 6, "x-end (m)", bold=True, align="right")
input_cell(ws2, 2, 7, 2000, "0")
label_cell(ws2, 2, 8, "Step (m)", bold=True, align="right")
input_cell(ws2, 2, 9, 20, "0")

label_cell(ws2, 3, 2, "x₀ pusat (m)", bold=True, align="right")
input_cell(ws2, 3, 3, 700, "0")

# ── ROW 5: separator + INPUT TABLE title ─────────────────────────────────────
ws2.row_dimensions[4].height = 6
ws2.merge_cells("A5:S5")
style(ws2["A5"], "▼  TABEL INPUT PARAMETER VARIASI  —  ubah nilai biru sesuai kebutuhan  ▼",
      bold=True, color=WHITE, bg="375623", sz=11)
ws2.row_dimensions[5].height = 24

# ── INPUT TABLE: VARIASI KEDALAMAN (rows 6-12) ───────────────────────────────
section_label(ws2, 6, 1, "GRAFIK 1 — Variasi Kedalaman z", bg="2E75B6", colspan=7)
ws2.row_dimensions[6].height = 22

# Header row 7
for col, txt in [(1,"Kurva"),(2,"z (m)"),(3,"ρ (kg/m³)"),(4,"R (m)"),
                 (5,"x₀ (m)"),(6,"Label Grafik"),(7,"Warna")]:
    header(ws2, 7, col, txt, bg="2E75B6")

# 5 input rows for z variation
z_data = [
    (1, 100, 470, 25, 700, "z = 100 m", "FF0000"),
    (2, 200, 470, 25, 700, "z = 200 m", "0070C0"),
    (3, 300, 470, 25, 700, "z = 300 m (ref)", "00B050"),
    (4, 400, 470, 25, 700, "z = 400 m", "FFC000"),
    (5, 500, 470, 25, 700, "z = 500 m", "7030A0"),
]
Z_INPUT_ROW = 8  # first data row of z input table
for i, (n, z, rho, R, x0, lbl, col_hex) in enumerate(z_data):
    row = Z_INPUT_ROW + i
    ws2.row_dimensions[row].height = 18
    label_cell(ws2, row, 1, f"Kurva {n}", align="center")
    for col, val in [(2,z),(3,rho),(4,R),(5,x0)]:
        input_cell(ws2, row, col, val, "0")
    c = ws2.cell(row, 6, lbl)
    c.font=Font(name="Arial",color=BLUE_INPUT,bold=True,size=10)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.fill=fill("EBF3FB"); c.border=all_border()
    c = ws2.cell(row, 7, col_hex)
    c.font=Font(name="Arial",color=BLUE_INPUT,bold=True,size=10)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.fill=fill("EBF3FB"); c.border=all_border()

# ── INPUT TABLE: VARIASI DENSITAS (rows 6-12, cols H+) ───────────────────────
section_label(ws2, 6, 9, "GRAFIK 2 — Variasi Densitas ρ", bg="833C11", colspan=5)
for col, txt in [(9,"Kurva"),(10,"ρ (kg/m³)"),(11,"z (m)"),(12,"R (m)"),(13,"x₀ (m)")]:
    header(ws2, 7, col, txt, bg="833C11")

rho_data = [
    (1, 200, 300, 25, 700),
    (2, 350, 300, 25, 700),
    (3, 470, 300, 25, 700),
    (4, 700, 300, 25, 700),
    (5, 900, 300, 25, 700),
]
RHO_INPUT_ROW = 8
for i, (n, rho, z, R, x0) in enumerate(rho_data):
    row = RHO_INPUT_ROW + i
    label_cell(ws2, row, 9, f"Kurva {n}", align="center")
    for col, val in [(10,rho),(11,z),(12,R),(13,x0)]:
        input_cell(ws2, row, col, val, "0")

# ── INPUT TABLE: VARIASI JARI-JARI (rows 6-12, cols N+) ─────────────────────
section_label(ws2, 6, 15, "GRAFIK 3 — Variasi Jari-Jari R", bg="1F4E79", colspan=5)
for col, txt in [(15,"Kurva"),(16,"R (m)"),(17,"z (m)"),(18,"ρ (kg/m³)"),(19,"x₀ (m)")]:
    header(ws2, 7, col, txt, bg="1F4E79")

R_data = [
    (1, 10,  300, 470, 700),
    (2, 15,  300, 470, 700),
    (3, 25,  300, 470, 700),
    (4, 40,  300, 470, 700),
    (5, 55,  300, 470, 700),
]
R_INPUT_ROW = 8
for i, (n, R, z, rho, x0) in enumerate(R_data):
    row = R_INPUT_ROW + i
    label_cell(ws2, row, 15, f"Kurva {n}", align="center")
    for col, val in [(16,R),(17,z),(18,rho),(19,x0)]:
        input_cell(ws2, row, col, val, "0")

# ── ROW 14: separator ─────────────────────────────────────────────────────────
ws2.row_dimensions[14].height = 6
ws2.merge_cells("A15:S15")
style(ws2["A15"], "▼  TABEL PERHITUNGAN  —  otomatis mengikuti input di atas  ▼",
      bold=True, color=WHITE, bg="1F4E79", sz=11)
ws2.row_dimensions[15].height = 24

# ── CALC TABLE HEADER row 16 ──────────────────────────────────────────────────
# Section colour bars row 16
section_label(ws2, 16, 3, "GRAFIK 1 — Variasi z", bg="2E75B6", colspan=5)
section_label(ws2, 16, 9, "GRAFIK 2 — Variasi ρ", bg="833C11", colspan=5)
section_label(ws2, 16, 15, "GRAFIK 3 — Variasi R", bg="1F4E79", colspan=5)

# Column headers row 17
# Formulas will reference input table labels
HDR_BG_Z   = "2E75B6"
HDR_BG_RHO = "833C11"
HDR_BG_R   = "1F4E79"
header(ws2, 17, 1, "No.", bg="595959")
header(ws2, 17, 2, "x (m)", bg="595959")
for i in range(5):
    # z headers pull label from input table
    c = ws2.cell(17, 3+i)
    c.value = f"=F{Z_INPUT_ROW+i}"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = fill(HDR_BG_Z)
    c.alignment = Alignment(horizontal="center", vertical="center")
    # rho headers
    c2 = ws2.cell(17, 9+i)
    c2.value = f"=CONCATENATE(\"\u03c1=\",TEXT({chr(ord('J'))+str(RHO_INPUT_ROW+i)},\"0\"),\" kg/m\u00b3\")"
    c2.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c2.fill = fill(HDR_BG_RHO)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    # R headers
    c3 = ws2.cell(17, 15+i)
    c3.value = f"=CONCATENATE(\"R=\",TEXT({chr(ord('P'))+str(R_INPUT_ROW+i)},\"0\"),\" m\")"
    c3.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c3.fill = fill(HDR_BG_R)
    c3.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[17].height = 22

# ── CALC DATA ROWS (rows 18 onward) ───────────────────────────────────────────
# x is generated from input: x_start, x_end, step
# x = x_start + (row_index)*step  but we cap at x_end
# Use 101 rows to cover 0..2000 at step 20 by default (user can change step)
CALC_START_2 = 18
N_CALC = 101

for i in range(N_CALC):
    row = CALC_START_2 + i
    ws2.row_dimensions[row].height = 16

    # No. and x column
    c_no = ws2.cell(row, 1, i)
    c_no.font = Font(name="Arial", size=9, color="595959")
    c_no.alignment = Alignment(horizontal="center", vertical="center")

    # x = x_start + i*step
    c_x = ws2.cell(row, 2)
    c_x.value = f"=$E$2+({i})*$I$2"
    c_x.font = Font(name="Arial", size=10, color=BLACK_CALC)
    c_x.alignment = Alignment(horizontal="right", vertical="center")
    c_x.number_format = "0"

    alt = (i % 2 == 0)
    alt_z   = "EBF3FB" if alt else "FFFFFF"
    alt_rho = "FEF4EE" if alt else "FFFFFF"
    alt_R   = "F1F8EC" if alt else "FFFFFF"

    # ── Grafik 1: Variasi z ───────────────────────────────────────────────
    # gz = G*(4π*ρ*R³) / (3*z²*(1+((x-x0)/z)²)^1.5)
    for j in range(5):
        col = 3 + j
        ir = Z_INPUT_ROW + j   # input row for this curve
        # refs from input table: z=B{ir}, rho=C{ir}, R=D{ir}, x0=E{ir}
        f = (f"={G_CELL}*(4*PI()*C{ir}*D{ir}^3)"
             f"/(3*B{ir}^2*(1+((B{row}-E{ir})^2/B{ir}^2))^1.5)")
        c = formula_cell(ws2, row, col, f)
        c.fill = fill(alt_z)

    # ── Grafik 2: Variasi ρ ───────────────────────────────────────────────
    # refs from rho input: rho=J{ir}, z=K{ir}, R=L{ir}, x0=M{ir}
    for j in range(5):
        col = 9 + j
        ir = RHO_INPUT_ROW + j
        f = (f"={G_CELL}*(4*PI()*J{ir}*L{ir}^3)"
             f"/(3*K{ir}^2*(1+((B{row}-M{ir})^2/K{ir}^2))^1.5)")
        c = formula_cell(ws2, row, col, f)
        c.fill = fill(alt_rho)

    # ── Grafik 3: Variasi R ───────────────────────────────────────────────
    # refs from R input: R=P{ir}, z=Q{ir}, rho=R{ir}, x0=S{ir}
    for j in range(5):
        col = 15 + j
        ir = R_INPUT_ROW + j
        f = (f"={G_CELL}*(4*PI()*R{ir}*P{ir}^3)"
             f"/(3*Q{ir}^2*(1+((B{row}-S{ir})^2/Q{ir}^2))^1.5)")
        c = formula_cell(ws2, row, col, f)
        c.fill = fill(alt_R)

# ── Grafik 1, 2, 3 charts ────────────────────────────────────────────────────
CHART_ROW_2 = f"A{CALC_START_2 + N_CALC + 2}"
make_chart(ws2, "Gambar 1. Efek Variasi Kedalaman Bola terhadap Anomali Gravitasi",
           2, [3,4,5,6,7],
           [f"=F{Z_INPUT_ROW+i}" for i in range(5)],
           CHART_ROW_2, CALC_START_2, N_CALC)

make_chart(ws2, "Gambar 2. Efek Variasi Densitas Bola terhadap Anomali Gravitasi",
           2, [9,10,11,12,13],
           [f"ρ={rho_data[i][1]} kg/m³" for i in range(5)],
           f"L{CALC_START_2 + N_CALC + 2}", CALC_START_2, N_CALC)

make_chart(ws2, "Gambar 3. Efek Variasi Jari-Jari Bola terhadap Anomali Gravitasi",
           2, [15,16,17,18,19],
           [f"R={R_data[i][1]} m" for i in range(5)],
           f"A{CALC_START_2 + N_CALC + 33}", CALC_START_2, N_CALC)

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 3  —  MODEL TABUNG 2.75D
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Sheet3_Tabung2.75D")
ws3.sheet_view.showGridLines = True

col_w3 = {"A":6,"B":12,"C":16,"D":16,"E":16,"F":16,"G":16,
           "H":4,"I":16,"J":16,"K":16,"L":16,"M":16,
           "N":4,"O":16,"P":16,"Q":16,"R":16,"S":16}
for c,w in col_w3.items(): ws3.column_dimensions[c].width = w

# ── Title ─────────────────────────────────────────────────────────────────────
ws3.merge_cells("A1:S1")
style(ws3["A1"], "MODEL TABUNG 2.75D — Pemodelan Gravitasi 2D",
      bold=True, color=WHITE, bg="375623", sz=14)
ws3.row_dimensions[1].height = 30

# ── Constant parameters rows 2-3 ─────────────────────────────────────────────
ws3.merge_cells("A2:A3"); style(ws3["A2"], "Parameter\nTetap", bold=True, bg="D6E4F0", sz=9)
label_cell(ws3, 2, 2, "G (mGal konst.)", bold=True, align="right")
cg = ws3.cell(2, 3, 2.097e-5)
cg.font=Font(name="Arial",color=BLUE_INPUT,bold=True,size=10)
cg.number_format="0.000E+00"
cg.alignment=Alignment(horizontal="center",vertical="center")
cg.fill=fill("EBF3FB"); cg.border=all_border()
G3_CELL = "$C$2"

label_cell(ws3, 2, 4, "x-start (m)", bold=True, align="right")
input_cell(ws3, 2, 5, 0, "0")
label_cell(ws3, 2, 6, "x-end (m)", bold=True, align="right")
input_cell(ws3, 2, 7, 2000, "0")
label_cell(ws3, 2, 8, "Step (m)", bold=True, align="right")
input_cell(ws3, 2, 9, 20, "0")

# ── Input table separator ─────────────────────────────────────────────────────
ws3.row_dimensions[4].height = 6
ws3.merge_cells("A5:S5")
style(ws3["A5"], "▼  TABEL INPUT PARAMETER VARIASI  —  ubah nilai biru sesuai kebutuhan  ▼",
      bold=True, color=WHITE, bg="375623", sz=11)
ws3.row_dimensions[5].height = 24

# ── GRAFIK 4 — Variasi Panjang Tabung (yb − ya) ──────────────────────────────
# REVISI: panjang = yb - ya (bukan ya=yb=Y)
section_label(ws3, 6, 1, "GRAFIK 4 — Variasi Panjang Tabung  [Panjang = yb − ya]", bg="375623", colspan=8)
ws3.row_dimensions[6].height = 22

for col, txt in [(1,"Kurva"),(2,"ya (m)"),(3,"yb (m)"),(4,"Panjang\n(yb−ya) m"),
                 (5,"ρ (kg/m³)"),(6,"R (m)"),(7,"z (m)"),(8,"x₀ (m)")]:
    header(ws3, 7, col, txt, bg="375623")

# Panjang = yb − ya; ya and yb are independent
g4_data = [
    (1, -100,  100, 470, 25, 300, 700),  # panjang 200
    (2, -250,  250, 470, 25, 300, 700),  # panjang 500
    (3, -500,  500, 470, 25, 300, 700),  # panjang 1000 (ref)
    (4,-1000, 1000, 470, 25, 300, 700),  # panjang 2000
    (5,-2000, 2000, 470, 25, 300, 700),  # panjang 4000
]
G4_INPUT_ROW = 8
for i, (n, ya, yb, rho, R, z, x0) in enumerate(g4_data):
    row = G4_INPUT_ROW + i
    ws3.row_dimensions[row].height = 18
    label_cell(ws3, row, 1, f"Kurva {n}", align="center")
    for col, val in [(2,ya),(3,yb),(5,rho),(6,R),(7,z),(8,x0)]:
        input_cell(ws3, row, col, val, "0")
    # Col 4: Panjang auto = yb - ya (formula, shown for info)
    c = ws3.cell(row, 4)
    c.value = f"=C{row}-B{row}"
    c.font = Font(name="Arial", color=BLACK_CALC, size=10, italic=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0"
    c.fill = fill("F2F2F2")
    c.border = all_border()

# ── GRAFIK 5 — Variasi Posisi Titik Tengah x₀ ────────────────────────────────
section_label(ws3, 6, 10, "GRAFIK 5 — Variasi Posisi Titik Tengah  x₀", bg="833C11", colspan=5)
for col, txt in [(10,"Kurva"),(11,"x₀ (m)"),(12,"ya (m)"),(13,"yb (m)"),
                 (14,"ρ (kg/m³)"),(15,"R (m)"),(16,"z (m)")]:
    header(ws3, 7, col, txt, bg="833C11")
# Actually 7 cols: K..Q = 11..17
# Re-do: 10=blank, 11..17 = 7 params
section_label(ws3, 6, 11, "GRAFIK 5 — Variasi Posisi Titik Tengah  x₀", bg="833C11", colspan=7)
for col, txt in [(11,"Kurva"),(12,"x₀ (m)"),(13,"ya (m)"),(14,"yb (m)"),
                 (15,"ρ (kg/m³)"),(16,"R (m)"),(17,"z (m)")]:
    header(ws3, 7, col, txt, bg="833C11")

g5_data = [
    (1,  300, -500,  500, 470, 25, 300),
    (2,  500, -500,  500, 470, 25, 300),
    (3,  700, -500,  500, 470, 25, 300),
    (4,  900, -500,  500, 470, 25, 300),
    (5, 1100, -500,  500, 470, 25, 300),
]
G5_INPUT_ROW = 8
for i, (n, x0, ya, yb, rho, R, z) in enumerate(g5_data):
    row = G5_INPUT_ROW + i
    label_cell(ws3, row, 11, f"Kurva {n}", align="center")
    for col, val in [(12,x0),(13,ya),(14,yb),(15,rho),(16,R),(17,z)]:
        input_cell(ws3, row, col, val, "0")

# ── GRAFIK 6 — Variasi y'b ────────────────────────────────────────────────────
section_label(ws3, 6, 19, "GRAFIK 6 — Variasi Posisi y'b  (ya tetap)", bg="1F4E79", colspan=6)
for col, txt in [(19,"Kurva"),(20,"ya (m)"),(21,"yb (m)"),(22,"ρ (kg/m³)"),
                 (23,"R (m)"),(24,"z (m)"),(25,"x₀ (m)")]:
    header(ws3, 7, col, txt, bg="1F4E79")
for c, w in [("T",16),("U",16),("V",16),("W",16),("X",16),("Y",16)]:
    ws3.column_dimensions[c].width = w

g6_data = [
    (1, -500,  100, 470, 25, 300, 700),
    (2, -500,  300, 470, 25, 300, 700),
    (3, -500,  500, 470, 25, 300, 700),
    (4, -500, 1000, 470, 25, 300, 700),
    (5, -500, 2000, 470, 25, 300, 700),
]
G6_INPUT_ROW = 8
for i, (n, ya, yb, rho, R, z, x0) in enumerate(g6_data):
    row = G6_INPUT_ROW + i
    label_cell(ws3, row, 19, f"Kurva {n}", align="center")
    for col, val in [(20,ya),(21,yb),(22,rho),(23,R),(24,z),(25,x0)]:
        input_cell(ws3, row, col, val, "0")

# ── CALC TABLE SEPARATOR ──────────────────────────────────────────────────────
ws3.row_dimensions[14].height = 6
ws3.merge_cells("A15:Y15")
style(ws3["A15"], "▼  TABEL PERHITUNGAN  —  otomatis mengikuti input di atas  ▼",
      bold=True, color=WHITE, bg="375623", sz=11)
ws3.row_dimensions[15].height = 24

# Section bars row 16
section_label(ws3, 16, 3, "GRAFIK 4 — Variasi Panjang (yb − ya)", bg="375623", colspan=5)
section_label(ws3, 16, 9, "GRAFIK 5 — Variasi x₀", bg="833C11", colspan=5)
section_label(ws3, 16, 15, "GRAFIK 6 — Variasi y'b", bg="1F4E79", colspan=5)

# Column headers row 17
header(ws3, 17, 1, "No.", bg="595959")
header(ws3, 17, 2, "x (m)", bg="595959")
for j in range(5):
    # G4 header: "Panjang = D{ir}" auto-label
    ir4 = G4_INPUT_ROW + j
    c = ws3.cell(17, 3+j)
    c.value = f"=CONCATENATE(\"Panjang \",TEXT(C{ir4}-B{ir4},\"0\"),\" m\")"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = fill("375623")
    c.alignment = Alignment(horizontal="center", vertical="center")
    # G5 header: "x₀=..."
    ir5 = G5_INPUT_ROW + j
    c2 = ws3.cell(17, 9+j)
    c2.value = f"=CONCATENATE(\"x\u2080=\",TEXT(L{ir5},\"0\"),\" m\")"
    c2.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c2.fill = fill("833C11")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    # G6 header: "yb=..."
    ir6 = G6_INPUT_ROW + j
    c3 = ws3.cell(17, 15+j)
    c3.value = f"=CONCATENATE(\"yb=\",TEXT(U{ir6},\"0\"),\" m\")"
    c3.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c3.fill = fill("1F4E79")
    c3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[17].height = 22

# ── CALC DATA ROWS ────────────────────────────────────────────────────────────
# gz = G_konst * ρ * R² / r² * (1/√(1+(r/ya)²) + 1/√(1+(r/yb)²))
# r = √((x - x0)² + z²)
# REVISI G4: ya and yb from input independently (panjang = yb - ya)

CALC_START_3 = 18
N3 = 101
ws3.freeze_panes = "B18"

for i in range(N3):
    row = CALC_START_3 + i
    ws3.row_dimensions[row].height = 16
    ws3.cell(row, 1, i).font = Font(name="Arial", size=9, color="595959")
    ws3.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")

    cx = ws3.cell(row, 2)
    cx.value = f"=$E$2+({i})*$I$2"
    cx.font = Font(name="Arial", size=10)
    cx.alignment = Alignment(horizontal="right", vertical="center")
    cx.number_format = "0"

    alt = (i % 2 == 0)
    alt_g4 = "EEFAE8" if alt else "FFFFFF"
    alt_g5 = "FEF4EE" if alt else "FFFFFF"
    alt_g6 = "EBF3FB" if alt else "FFFFFF"

    # ── Grafik 4: Variasi Panjang (yb−ya), REVISI ─────────────────────────
    for j in range(5):
        ir = G4_INPUT_ROW + j
        col = 3 + j
        # ya = B{ir}, yb = C{ir}, rho = E{ir}, R = F{ir}, z = G{ir}, x0 = H{ir}
        # r = sqrt((x - x0)^2 + z^2)
        r_expr = f"SQRT((B{row}-H{ir})^2+G{ir}^2)"
        f = (f"=IF(({r_expr})=0,0,"
             f"{G3_CELL}*E{ir}*F{ir}^2/({r_expr})^2"
             f"*(1/SQRT(1+({r_expr}/ABS(B{ir}))^2)"
             f"+1/SQRT(1+({r_expr}/ABS(C{ir}))^2)))")
        c = formula_cell(ws3, row, col, f)
        c.fill = fill(alt_g4)

    # ── Grafik 5: Variasi x₀ ──────────────────────────────────────────────
    for j in range(5):
        ir = G5_INPUT_ROW + j
        col = 9 + j
        # x0=L{ir}, ya=M{ir}, yb=N{ir}, rho=O{ir}, R=P{ir}, z=Q{ir}
        r_expr = f"SQRT((B{row}-L{ir})^2+Q{ir}^2)"
        f = (f"=IF(({r_expr})=0,0,"
             f"{G3_CELL}*O{ir}*P{ir}^2/({r_expr})^2"
             f"*(1/SQRT(1+({r_expr}/ABS(M{ir}))^2)"
             f"+1/SQRT(1+({r_expr}/ABS(N{ir}))^2)))")
        c = formula_cell(ws3, row, col, f)
        c.fill = fill(alt_g5)

    # ── Grafik 6: Variasi y'b ──────────────────────────────────────────────
    for j in range(5):
        ir = G6_INPUT_ROW + j
        col = 15 + j
        # ya=T{ir}, yb=U{ir}, rho=V{ir}, R=W{ir}, z=X{ir}, x0=Y{ir}
        r_expr = f"SQRT((B{row}-Y{ir})^2+X{ir}^2)"
        f = (f"=IF(({r_expr})=0,0,"
             f"{G3_CELL}*V{ir}*W{ir}^2/({r_expr})^2"
             f"*(1/SQRT(1+({r_expr}/ABS(T{ir}))^2)"
             f"+1/SQRT(1+({r_expr}/ABS(U{ir}))^2)))")
        c = formula_cell(ws3, row, col, f)
        c.fill = fill(alt_g6)

# ── Charts Sheet3 ─────────────────────────────────────────────────────────────
CR3 = CALC_START_3 + N3 + 2
make_chart(ws3, "Gambar 4. Efek Variasi Panjang Tabung 2.75D [Panjang = yb − ya]",
           2, [3,4,5,6,7],
           [f"=CONCATENATE(\"Panjang \",TEXT(C{G4_INPUT_ROW+i}-B{G4_INPUT_ROW+i},\"0\"),\" m\")"
            for i in range(5)],
           f"A{CR3}", CALC_START_3, N3)

make_chart(ws3, "Gambar 5. Efek Variasi Posisi Titik Tengah x₀",
           2, [9,10,11,12,13],
           [f"x₀={g5_data[i][1]} m" for i in range(5)],
           f"L{CR3}", CALC_START_3, N3)

make_chart(ws3, "Gambar 6. Efek Variasi Posisi y'b (ya tetap)",
           2, [15,16,17,18,19],
           [f"yb={g6_data[i][2]} m (ya={g6_data[i][1]} m)" for i in range(5)],
           f"A{CR3+31}", CALC_START_3, N3)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save("/home/claude/PDG_W8_Flexible.xlsx")
print("Saved OK")